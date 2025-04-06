

from eve_api_requests import EveRequester
from esi_IDLoader import esi_ID_loader
from datetime import datetime, timedelta, date
from openpyxl.styles import Font, numbers  # Добавляем импорт
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import json
import openpyxl
import time

def check_field(json, field_name):
    if field_name in json:
        return json[field_name]
    else:
        print(f"Поля `{field_name}` нет в JSON.")

def main():
    loader = esi_ID_loader()
    current_date = datetime.now()
    last_month = date.today().replace(day=1) - timedelta(days=1)
    month = last_month.month
    year = last_month.year
    print(f"month {month} year {year}")
    print(f"{loader.cta_systems}")
    requester = EveRequester()
    inshu = requester.get_Inshurances()
    inshurances = dict()
    for ins in inshu:
        for level in ins["levels"]:
            if level["name"] == "Platinum":
                inshurances[ins["type_id"]] = level["payout"] - level["cost"]

    Members = dict()
    for aliance in loader.alliances:
        AliInfo = requester.get_Aliance_Info(aliance)
        AlianceTicker = AliInfo["ticker"]
        for p in range(1,15):
            aliance_killmails = requester.get_zkillboard_killmails_for_alliance(aliance,year,month,p)
            for km in aliance_killmails:
                killmail_id = km["killmail_id"]
                killmail_hash = km["zkb"]["hash"]
                totalValue = km["zkb"]["fittedValue"]
                cost = 0
                if not km["zkb"]["npc"]:
                    kmd = requester.get_killmail_details(killmail_id, killmail_hash)
                    if "victim" in kmd:
                        if "character_id" in kmd["victim"]:
                            char_id = kmd["victim"]["character_id"]
                        else:
                            print(f"Запись о структуре {killmail_id} пропущу это килмыло.")
                            continue
                    else:
                        print(f"Запись о структуре {killmail_id} пропущу это килмыло.")
                        continue
                    systemid = kmd["solar_system_id"]
                    if len(loader.cta_systems) > 0:
                        if str(systemid) not in loader.cta_systems :
                            print(f"{systemid} Система не заапрувлена как кта. Wrong solar system. It will be passed")
                            continue

                    if char_id not in Members:
                        char_info = requester.get_CharacterInfo(kmd["victim"]["character_id"])
                        if str(char_info.get('alliance_id', '')) in loader.alliances:
                            new_member = {
                                "ID": char_id,
                                "Name": char_info["name"],
                                "Aliance": AlianceTicker,
                                "total_cost": 0,
                                "natur_compens_ids" : [],
                                "dps_links": [],
                                "support_links": [],
                                "tackle_links": [],
                                "valuble_links": [],
                                "capital_links": [],
                                "LightShip_links": []
                            }
                            Members[new_member["ID"]] = new_member
                        else:
                            continue

                    sh_id = kmd["victim"]["ship_type_id"]
                    ship_id = str(sh_id)
                    link = f"https://zkillboard.com/kill/{killmail_id}/"
                    cost = 0
                    ktime = kmd["killmail_time"]
                    # forbidden = [33475, 33700]
                    if sh_id in inshurances:
                        cost = totalValue - inshurances[sh_id]

                    if ship_id in loader.support_ships:
                        Members[char_id]["total_cost"] += cost
                        Members[char_id]["support_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " Большой молодец ")
                    elif ship_id in loader.tackle_ships:
                        Members[char_id]["total_cost"] += cost
                        Members[char_id]["tackle_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " молодец ")
                    elif ship_id in loader.dps_ships:
                        Members[char_id]["total_cost"] += cost * 0.7
                        Members[char_id]["dps_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " норм ")
                    elif ship_id in loader.vaiuble_ships:
                        Members[char_id]["natur_compens_ids"].append(ship_id)
                        Members[char_id]["total_cost"] -= inshurances[sh_id]
                        cost = -inshurances[sh_id]
                        Members[char_id]["valuble_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " молодец ")
                    elif ship_id in loader.capital_ships:
                        Members[char_id]["natur_compens_ids"].append(ship_id)
                        Members[char_id]["total_cost"] -= inshurances[sh_id]
                        cost = -inshurances[sh_id]
                        Members[char_id]["capital_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " Вообще молодец ")
                    elif ship_id in loader.t3_ships:
                        items = kmd["victim"]["items"]
                        has_valid_item = any(
                            str(item.get("item_type_id")) == "45609" and str(item.get("flag")) != "5"
                            for item in items
                        )
                        if has_valid_item:
                            Members[char_id]["total_cost"] += cost
                            Members[char_id]["support_links"].append({"cost": cost, "link": link, "ktime": ktime})
                            print(Members[char_id]["Name"] + " Большой молодец ")
                        else:
                            Members[char_id]["total_cost"] += cost * 0.7
                            Members[char_id]["dps_links"].append({"cost": cost, "link": link, "ktime": ktime})
                            print(Members[char_id]["Name"] + " норм ")
                    elif ship_id in loader.light_ships:
                        Members[char_id]["LightShip_links"].append({"cost": cost, "link": link, "ktime": ktime})
                        print(Members[char_id]["Name"] + " водит вигилы")
                    else:
                        print(Members[char_id]["Name"] + " крабина ")
                else:
                    print("Килл от нпц")
            time.sleep(2)

    MData = dict()
    for char_id, member_data in Members.items():
        if member_data["total_cost"] != 0:
            main_data = {
                "ID": member_data["ID"],
                "Name": member_data["Name"],
                "Aliance": member_data["Aliance"],
                "total_cost": member_data["total_cost"],
                "natur_compens_ids": member_data["natur_compens_ids"]
            }
            MData[char_id] = main_data

    main_df = pd.DataFrame(list(MData.values()))
    links_data = []
    # Проходимся по каждому члену в Members
    for char_id, member_data in Members.items():
        # Проверяем, что total_cost != 0
        if member_data["total_cost"] != 0:
            # Проходимся по всем типам линков
            for link_type, links in member_data.items():
                if "links" in link_type and links:  # Проверяем, что это список линков и он не пустой
                    for link in links:
                        # Добавляем данные в links_data
                        links_data.append({
                            "Name": member_data["Name"],
                            "Aliance": member_data["Aliance"],
                            "Link Type": link_type,
                            "Cost": link["cost"],
                            "ktime": link["ktime"],
                            "Link": link["link"]
                        })

    links_df = pd.DataFrame(links_data)
    # Сортируем DataFrame по столбцу "Name" (по алфавиту)
    links_df = links_df.sort_values(by=["Aliance", "Name"])

    # # Фильтрация словаря
    # filtered_data = {k: v for k, v in Members.items()}
    #
    # # Исключаем из словаря все элементы, у которых total_cost и natur_compens_ids пустые
    # filtered_data = {k: v for k, v in filtered_data.items() if v["total_cost"] > 0 or len(v["natur_compens_ids"]) > 0}
    #
    # df = pd.DataFrame(list(filtered_data.values()))
    timestamp = current_date.strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"output_{timestamp}.xlsx"
    # df.to_excel(filename, index=False)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 1. Сортируем данные
        main_df = main_df.sort_values(by=['Aliance', 'Name'])
        links_df = links_df.sort_values(by=['Aliance', 'Name'])

        # 2. Создаем summary_df (без total_cost)
        summary_df = main_df[['Name', 'Aliance']].drop_duplicates()
        summary_df['Formation ships sum'] = 0
        summary_df['Small ships sum'] = 0
        summary_df['Total sum'] = 0

        # 3. Записываем данные
        main_df.to_excel(writer, sheet_name="Main", index=False)  # Пока с исходными значениями
        links_df.to_excel(writer, sheet_name="Links for check", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # 4. Получаем доступ к листам
        workbook = writer.book
        ws_main = writer.sheets['Main']
        ws_summary = writer.sheets['Summary']
        ws_links = writer.sheets['Links for check']

        # 5. Добавляем формулы в Summary
        for i in range(2, len(summary_df) + 2):
            # Суммы по типам кораблей
            ws_summary[
                f'C{i}'] = f'=SUMIFS(\'{ws_links.title}\'!D:D, \'{ws_links.title}\'!A:A, A{i}, \'{ws_links.title}\'!C:C, "<>LightShip_links")'
            ws_summary[
                f'D{i}'] = f'=SUMIFS(\'{ws_links.title}\'!D:D, \'{ws_links.title}\'!A:A, A{i}, \'{ws_links.title}\'!C:C, "LightShip_links")'
            ws_summary[f'E{i}'] = f'=C{i}+D{i}'

        # 6. Заменяем значения в Main на формулы
        total_cost_col = main_df.columns.get_loc('total_cost') + 1  # Номер колонки total_cost

        for main_row in range(2, len(main_df) + 2):
            name = ws_main[f'B{main_row}'].value
            # Ищем соответствующую строку в Summary
            for sum_row in range(2, len(summary_df) + 2):
                if ws_summary[f'A{sum_row}'].value == name:
                    # Заменяем значение на формулу
                    ws_main.cell(
                        row=main_row,
                        column=total_cost_col,
                        value=f'=Summary!E{sum_row}'  # Ссылка на Total sum
                    )
                    break

        # 7. Форматирование
        for sheet in [ws_main, ws_summary]:
            # Жирные заголовки
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            # Числовой формат для сумм
            for row in range(2, len(summary_df) + 2):
                if sheet == ws_main:
                    ws_main.cell(row=row, column=total_cost_col).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    for col in ['C', 'D', 'E']:
                        ws_summary.cell(row=row,
                                        column=ord(col) - 64).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # 8. Итоговая строка в Summary
        last_row = len(summary_df) + 2
        ws_summary[f'B{last_row}'] = 'TOTAL'
        ws_summary[f'C{last_row}'] = f'=SUM(C2:C{last_row - 1})'
        ws_summary[f'D{last_row}'] = f'=SUM(D2:D{last_row - 1})'
        ws_summary[f'E{last_row}'] = f'=SUM(E2:E{last_row - 1})'
        for col in ['B', 'C', 'D', 'E']:
            ws_summary[f'{col}{last_row}'].font = Font(bold=True)

    print(f"Excel-файл '{filename}' успешно создан!")



if __name__ == "__main__":
    main()